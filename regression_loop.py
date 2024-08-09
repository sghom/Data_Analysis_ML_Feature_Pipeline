import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
import statsmodels.api as sm
from statsmodels.stats.outliers_influence import variance_inflation_factor
from sklearn.linear_model import LogisticRegression, Lasso, ElasticNetCV
from sklearn.feature_selection import SelectFromModel
from sklearn.ensemble import ExtraTreesClassifier, GradientBoostingClassifier, RandomForestClassifier
from sklearn.model_selection import train_test_split
from sklearn.metrics import roc_curve, auc
from matplotlib.font_manager import FontProperties
import os
from openpyxl import Workbook, load_workbook

# Define the path to your Excel file
path = 'Z:/ra/uhn/blad/Meeting - Aug 16th/regression.xlsx'
sheet_names = ['Clinical','qCT','PFT','Osc','Clinical_PFT','Clinical_Osc',
               'Clinical_qCT','PFT_Osc','PFT_qCT','Osc_qCT','Clinical_PFT_Osc',
               'Clinical_PFT_qCT','Clinical_Osc_qCT','PFT_Osc_qCT',
               'Clinical_PFT_Osc_qCT']  # Add your sheet names here

for sheet_name in sheet_names:
    # Read the specified columns from the Excel sheet
    reg_df = pd.read_excel(path, sheet_name=sheet_name)
    reg_df = reg_df.dropna(axis=1)

    dependent_var = 'blad_status'
    covariates = reg_df.columns.difference([dependent_var])

    # Correlation Analysis
    correlation_matrix = reg_df[[dependent_var] + list(covariates)].corr()

    # Find highly correlated pairs
    threshold = 0.9  # Adjust as needed
    highly_correlated_pairs = {}
    for i in range(len(correlation_matrix.columns)):
        for j in range(i+1, len(correlation_matrix.columns)):
            if abs(correlation_matrix.iloc[i, j]) > threshold:
                colname_i = correlation_matrix.columns[i]
                colname_j = correlation_matrix.columns[j]
                pair_name = f"{colname_i} - {colname_j}"
                highly_correlated_pairs[pair_name] = (colname_i, colname_j)

    # Remove highly correlated pairs from covariates list
    for var1, var2 in highly_correlated_pairs.values():
        if var1 in covariates:
            covariates = covariates.difference([var2])
        if var2 in covariates:
            covariates = covariates.difference([var1])

    # Function to calculate VIF
    def calculate_vif(df, features):
        X = df[features]
        X = sm.add_constant(X)
        vif = pd.DataFrame()
        vif["Variable"] = X.columns
        vif["VIF"] = [variance_inflation_factor(X.values, i) for i in range(X.shape[1])]
        return vif

    # Calculate VIF for updated covariates list
    vif_df = calculate_vif(reg_df, list(covariates))

    # Drop covariates with VIF greater than 5
    high_vif_vars = vif_df[vif_df["VIF"] > 5]["Variable"].tolist()

    # Remove high VIF variables from covariates list
    filtered_covariates = [var for var in covariates if var not in high_vif_vars]

    # Recalculate VIF for filtered covariates
    filtered_vif_df = calculate_vif(reg_df, filtered_covariates)

    print(f"\nUpdated Covariates List and their VIF after removing high VIF variables for sheet '{sheet_name}':")
    print(filtered_vif_df)

    X = reg_df[filtered_covariates]
    y = reg_df[dependent_var]
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.25, random_state=1)

    odds = (len(np.array(y)) - sum(np.array(y))) / sum(np.array(y))
    weights = {0: odds, 1: 1}

    # Logistic Regression
    model_lr = LogisticRegression(penalty='l2', max_iter=100000, random_state=0, class_weight=weights)
    model_lr.fit(X_train, y_train)

    # Get coefficients and store in a dataframe
    coef_df = pd.DataFrame({'Variable': X.columns, 'Coefficient': model_lr.coef_[0]})

    # Example with ExtraTreesClassifier for feature selection
    extra_trees = ExtraTreesClassifier(n_estimators=100, random_state=0, class_weight=weights)
    extra_trees.fit(X_train, y_train)

    # Select features based on importance
    sfm_extra_trees = SelectFromModel(extra_trees, prefit=True)
    selected_features_extra_trees = X_train.columns[sfm_extra_trees.get_support()]

    # Example with ElasticNetCV for feature selection
    elastic_net = ElasticNetCV(l1_ratio=0.5, cv=10, random_state=0)
    elastic_net.fit(X_train, y_train)

    # Select features based on coefficients
    coef_mask = elastic_net.coef_ != 0
    selected_features_elastic_net = X_train.columns[coef_mask]

    # Example with GradientBoostingClassifier for feature selection
    gb_classifier = GradientBoostingClassifier(n_estimators=100, learning_rate=1.0, max_depth=1, random_state=0)
    gb_classifier.fit(X_train, y_train)

    # Select features based on importance
    sfm_gb = SelectFromModel(gb_classifier, prefit=True)
    selected_features_gb = X_train.columns[sfm_gb.get_support()]

    # Example with Lasso for feature selection
    lasso = Lasso(alpha=0.1, random_state=0)
    lasso.fit(X_train, y_train)

    # Select features based on coefficients
    lasso_mask = lasso.coef_ != 0
    selected_features_lasso = X_train.columns[lasso_mask]

    # Example with RandomForestClassifier for feature selection
    rf_classifier = RandomForestClassifier(n_estimators=100, random_state=0, class_weight=weights)
    rf_classifier.fit(X_train, y_train)

    # Select features based on importance
    sfm_rf = SelectFromModel(rf_classifier, prefit=True)
    selected_features_rf = X_train.columns[sfm_rf.get_support()]

    # Determine the maximum length of selected features among all models
    max_len = max(len(coef_df['Variable']),
                  len(selected_features_extra_trees),
                  len(selected_features_elastic_net),
                  len(selected_features_gb),
                  len(selected_features_lasso),
                  len(selected_features_rf))

    # Create a dictionary to hold selected features for each model
    selected_features_dict = {
        'Logistic Regression': list(coef_df['Variable']) + [''] * (max_len - len(coef_df['Variable'])),
        'Extra Trees': list(selected_features_extra_trees) + [''] * (max_len - len(selected_features_extra_trees)),
        'Elastic Net': list(selected_features_elastic_net) + [''] * (max_len - len(selected_features_elastic_net)),
        'Gradient Boosting': list(selected_features_gb) + [''] * (max_len - len(selected_features_gb)),
        'Lasso': list(selected_features_lasso) + [''] * (max_len - len(selected_features_lasso)),
        'Random Forest': list(selected_features_rf) + [''] * (max_len - len(selected_features_rf))
    }

    # Create DataFrame from dictionary
    selected_features_df = pd.DataFrame(selected_features_dict)

    # Save selected_features_df to a separate Excel file for each sheet
    save_excel_path = f'Z:/ra/uhn/blad/Meeting - Aug 16th/selected_features/_{sheet_name}.xlsx'
    selected_features_df.to_excel(save_excel_path, index=False)

    # Print selected features for each model
    print(f"\nSelected Features for Each Model for sheet '{sheet_name}':")
    print(selected_features_df)

    # Fit models and plot ROC curves
    models = {
        'Logistic Regression': model_lr,
        'Extra Trees': extra_trees,
        'Gradient Boosting': gb_classifier,
        'Elastic Net': elastic_net,
        'Lasso': lasso,
        'Random Forest': rf_classifier
    }

    plt.figure(figsize=(13, 10))

    for name, model in models.items():
        if name == 'Logistic Regression':
            y_pred_proba = model.predict_proba(X_test)[:, 1]
        elif name == 'Elastic Net':
            y_pred_proba = elastic_net.predict(X_test)  # For ElasticNetCV, predict gives the final model's predictions directly
        elif name == 'Lasso':
            y_pred_proba = lasso.predict(X_test)
        else:
            y_pred_proba = model.predict_proba(X_test)[:, 1]

        fpr, tpr, _ = roc_curve(y_test, y_pred_proba)
        roc_auc = auc(fpr, tpr)

        plt.plot(fpr, tpr, lw=4, label=f'{name} (AUC = {roc_auc:.2f})')

    font_prop = FontProperties()
    font_prop.set_family('Arial')
    font_prop.set_size('14')

    plt.plot([0, 1], [0, 1], color='black', lw=3, linestyle='--')
    plt.xlim([0.0, 1.0])
    plt.ylim([0.0, 1.0])
    plt.xlabel('False Positive Rate', fontproperties=font_prop, fontsize=18, fontweight='bold')
    plt.ylabel('True Positive Rate', fontproperties=font_prop, fontsize=18, fontweight='bold')
    plt.title(f'{sheet_name} - ROC Curve', fontproperties=font_prop, fontsize=18, fontweight='bold')
    plt.xticks(fontproperties=font_prop)
    plt.yticks(fontproperties=font_prop)
    plt.legend(loc="lower right", fontsize=14)
    
    # Save the plot to the specified path
    save_path = f'Z:/ra/uhn/blad/Meeting - Aug 16th/figures/roc_curves/{sheet_name}_ROC_curve.tif'
    plt.savefig(save_path, format='tif', bbox_inches='tight')
    
    # # Display the plot
    # plt.show()

# Define the path where the individual Excel files are saved
folder_path = 'Z:/ra/uhn/blad/Meeting - Aug 16th/selected_features'

# Define the path for the merged Excel file
merged_excel_path = 'Z:/ra/uhn/blad/Meeting - Aug 16th/selected_features.xlsx'

# List all files in the folder path
file_names = os.listdir(folder_path)

# Initialize a new workbook
merged_workbook = Workbook()

# Iterate through each file name
for file_name in file_names:
    # Check if the file is an Excel file
    if file_name.endswith('.xlsx'):
        # Extract the sheet name from the file name
        sheet_name = os.path.splitext(file_name)[0]

        # Load the existing Excel file
        excel_file = load_workbook(os.path.join(folder_path, file_name))

        # Iterate through each sheet in the loaded workbook
        for sheet in excel_file.sheetnames:
            # Copy each sheet to the merged workbook
            source_sheet = excel_file[sheet]
            target_sheet = merged_workbook.create_sheet(title=f"{sheet_name}")
            
            for row in source_sheet.iter_rows():
                for cell in row:
                    target_sheet[cell.coordinate].value = cell.value

# Save the merged workbook
merged_workbook.save(merged_excel_path)